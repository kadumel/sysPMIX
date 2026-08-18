import csv
import io
import re
import unicodedata
from typing import Any

from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction

from api_sankhya.models import Cliente as ClienteSankhya

from .models import PerfilUsuario, UsuarioClienteSankhya

HEADER_ALIASES = {
    'codigo_cliente': {
        'codigo_cliente',
        'codigo',
        'cod_cliente',
        'codcliente',
        'c_digo_cliente',
    },
    'username': {
        'usuario',
        'username',
        'login',
        'user',
    },
    'first_name': {
        'nome',
        'first_name',
        'name',
    },
    'email': {
        'e_mail_gerado',
        'email_gerado',
        'e_mail',
        'email',
    },
    'password': {
        'senha_gerada',
        'senha',
        'password',
    },
}

REQUIRED_FIELDS = ('codigo_cliente', 'username', 'password')


def _norm_header(value: Any) -> str:
    text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode()
    text = text.lower().strip()
    return re.sub(r'[^a-z0-9]+', '_', text).strip('_')


def _cell_str(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _codigo_cliente(value: Any) -> int | None:
    text = _cell_str(value)
    if not text:
        return None
    if text.endswith('.0') and text.replace('.', '', 1).isdigit():
        text = text[:-2]
    try:
        return int(text)
    except (TypeError, ValueError):
        return None


def _map_headers(headers: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        key = _norm_header(raw)
        if not key:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _rows_from_xlsx(uploaded) -> list[list[Any]]:
    from openpyxl import load_workbook

    uploaded.seek(0)
    wb = load_workbook(uploaded, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            if not rows:
                continue
            mapping = _map_headers(rows[0])
            if all(field in mapping for field in REQUIRED_FIELDS):
                return rows
        return rows if rows else []
    finally:
        wb.close()


def _rows_from_csv(uploaded) -> list[list[Any]]:
    uploaded.seek(0)
    raw = uploaded.read()
    if isinstance(raw, bytes):
        text = raw.decode('utf-8-sig', errors='replace')
    else:
        text = str(raw)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return [list(row) for row in reader]


def ler_linhas_planilha(uploaded) -> tuple[list[dict[str, Any]], list[str]]:
    """Lê xlsx/csv no formato Codigo_Cliente, Usuario, Nome, E-mail, Senha."""
    erros: list[str] = []
    name = (getattr(uploaded, 'name', '') or '').lower()
    try:
        if name.endswith('.csv'):
            rows = _rows_from_csv(uploaded)
        else:
            rows = _rows_from_xlsx(uploaded)
    except Exception as exc:
        return [], [f'Não foi possível ler a planilha: {exc}']

    if not rows:
        return [], ['A planilha está vazia.']

    mapping = _map_headers(rows[0])
    missing = [f for f in REQUIRED_FIELDS if f not in mapping]
    if missing:
        labels = {
            'codigo_cliente': 'Codigo_Cliente',
            'username': 'Usuario',
            'password': 'Senha Gerada',
        }
        nomes = ', '.join(labels[f] for f in missing)
        return [], [f'Colunas obrigatórias não encontradas: {nomes}.']

    linhas: list[dict[str, Any]] = []
    for num, row in enumerate(rows[1:], start=2):
        def col(field: str) -> Any:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        codigo = _codigo_cliente(col('codigo_cliente'))
        username = _cell_str(col('username'))
        password = _cell_str(col('password'))
        first_name = _cell_str(col('first_name'))[:150]
        email = _cell_str(col('email'))[:254]
        if not any([codigo, username, password, first_name, email]):
            continue
        linhas.append({
            'linha': num,
            'codigo_cliente': codigo,
            'username': username,
            'password': password,
            'first_name': first_name,
            'email': email,
        })
    if not linhas:
        erros.append('Nenhuma linha de dados encontrada na planilha.')
    return linhas, erros


def importar_usuarios_planilha(uploaded) -> dict[str, Any]:
    linhas, erros_leitura = ler_linhas_planilha(uploaded)
    resultado = {
        'criados': 0,
        'ignorados': 0,
        'erros': list(erros_leitura),
    }
    if erros_leitura:
        return resultado

    clientes = {
        c.codigo_cliente: c
        for c in ClienteSankhya.objects.all().only('id', 'codigo_cliente')
    }
    usernames_existentes = {
        u.lower() for u in User.objects.values_list('username', flat=True)
    }
    vistos_no_arquivo: set[str] = set()

    for item in linhas:
        linha = item['linha']
        codigo = item['codigo_cliente']
        username = item['username']
        password = item['password']

        if codigo is None:
            resultado['erros'].append(f'Linha {linha}: código do cliente inválido.')
            continue
        if not username:
            resultado['erros'].append(f'Linha {linha}: usuário (login) vazio.')
            continue
        if len(username) > 150:
            resultado['erros'].append(f'Linha {linha}: login com mais de 150 caracteres.')
            continue
        if not password:
            resultado['erros'].append(f'Linha {linha}: senha vazia.')
            continue

        cliente = clientes.get(codigo)
        if cliente is None:
            resultado['erros'].append(
                f'Linha {linha}: cliente código {codigo} não encontrado em sankhya_cliente.'
            )
            continue

        username_key = username.lower()
        if username_key in vistos_no_arquivo:
            resultado['erros'].append(f'Linha {linha}: login "{username}" duplicado na planilha.')
            continue
        if username_key in usernames_existentes:
            resultado['ignorados'] += 1
            continue

        try:
            validate_password(password, user=User(username=username, email=item['email']))
        except DjangoValidationError as exc:
            resultado['erros'].append(
                f'Linha {linha} ({username}): senha inválida — {"; ".join(exc.messages)}'
            )
            continue

        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    email=item['email'] or '',
                    password=password,
                    first_name=item['first_name'] or '',
                )
                PerfilUsuario.objects.filter(user=user).update(perfil=PerfilUsuario.Perfil.CLIENTE)
                UsuarioClienteSankhya.objects.create(cliente=cliente, user=user)
        except Exception as exc:
            resultado['erros'].append(f'Linha {linha} ({username}): {exc}')
            continue

        vistos_no_arquivo.add(username_key)
        usernames_existentes.add(username_key)
        resultado['criados'] += 1

    return resultado
